import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import dropbox
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

# --- 1. CONEXIÓN A DROPBOX ---
try:
    dbx = dropbox.Dropbox(st.secrets["DROPBOX_TOKEN"])
    # LA RUTA EXACTA DE TU CARPETA COMPARTIDA:
    DBX_PATH = "/Grupo Ejercicio Físico y Nutrición/App_Fuerza/Base_de_Datos_Fuerza.xlsx"
except Exception as e:
    st.error("⚠️ Falta configurar el token de Dropbox en Streamlit Secrets.")

# --- CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(page_title="NI-Force Cloud", layout="wide")
st.title("☁️ Analizador de Fuerza NI - Conectado a Dropbox")
st.info("Sube tus CSV. Los resultados se guardarán automáticamente en la base de datos central compartida del grupo.")

# --- SIDEBAR (DATOS Y AJUSTES) ---
st.sidebar.header("📋 Datos del Lote")
id_sujeto = st.sidebar.text_input("ID Sujeto (NIxx)", value="NI00", max_chars=4).strip().upper()
es_id_valido = bool(re.match(r"^NI\d{2}$", id_sujeto))

if not es_id_valido:
    st.sidebar.error("❌ Formato requerido: NI00")
else:
    st.sidebar.success(f"✅ ID Correcto")

nombre_sesion = st.sidebar.text_input("Sesión", "Sesión_A").strip()
serie_inicial = st.sidebar.number_input("Empezar en Serie nº", min_value=1, value=1)

st.sidebar.header("⚙️ Ajustes")
sens_inicio = st.sidebar.slider("Inicio (N/s)", 10, 2000, 40)
sens_final_pendiente = st.sidebar.slider("Final (Sensibilidad)", -1000, -5, -10)
recorte = st.sidebar.slider("Recorte (s)", 1.0, 10.0, 5.0)
umbral_corte = st.sidebar.number_input("Valor X final (N)", value=25.0)

COLORES_PASTEL = ["FFEBEE", "E3F2FD", "F1F8E9", "FFF3E0", "F3E5F5", "E0F7FA", "FFFDE7", "F9FBE7", "E8EAF6", "EFEBE9"]

def estilizar_excel(df, output):
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Resultados')
        ws = writer.sheets['Resultados']
        ref = f"A1:{chr(64 + len(df.columns))}{len(df) + 1}"
        tab = Table(displayName="DB_Fuerza", ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tab)
        
        ids = df['Sujeto'].unique().tolist()
        mapa = {id: COLORES_PASTEL[i % len(COLORES_PASTEL)] for i, id in enumerate(ids)}
        for row in range(2, ws.max_row + 1):
            color = mapa.get(ws.cell(row=row, column=2).value, "FFFFFF")
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=row, column=col).fill = fill
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
    return output

# --- PROCESAMIENTO ---
csv_files = st.file_uploader("📊 Sube los archivos CSV", type="csv", accept_multiple_files=True)

if csv_files:
    nuevas_reps = []
    
    for idx, arc in enumerate(csv_files):
        try:
            serie_act = serie_inicial + idx
            df_raw = pd.read_csv(arc, sep=';', header=None)
            t = pd.to_numeric(df_raw[0], errors='coerce').values / 1_000_000.0
            f = pd.to_numeric(df_raw[1].astype(str).str.replace(',', '.'), errors='coerce').values
            mask = ~np.isnan(t) & ~np.isnan(f)
            t, f = t[mask], f[mask]
            
            pend = np.gradient(f) / np.gradient(t)
            inicios, en_accion = [], False
            for i in range(1, len(pend)):
                if pend[i] > sens_inicio and not en_accion:
                    inicios.append(i); en_accion = True
                elif f[i] < (umbral_corte / 2) and en_accion:
                    en_accion = False
            
            for i, idx_ini in enumerate(inicios):
                t_c = t[idx_ini] + recorte
                idxs_f = np.where((t > t_c) & (f < umbral_corte))[0]
                if len(idxs_f) > 0:
                    idx_f = idxs_f[0]
                    for j in range(idx_f, idx_ini, -1):
                        if pend[j] > sens_final_pendiente:
                            idx_f = j; break
                    
                    m_m = (t >= t_c) & (t <= t[idx_f])
                    if len(f[m_m]) > 3:
                        nuevas_reps.append({
                            'Fecha': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M'),
                            'Sujeto': id_sujeto,
                            'Sesión': nombre_sesion,
                            'SERIE': int(serie_act),
                            'REPETICIÓN': i + 1,
                            'Media(N)': round(np.mean(f[m_m]), 2),
                            'Max(N)': round(np.max(f[m_m]), 2)
                        })
        except Exception: pass

    if nuevas_reps:
        df_nuevas = pd.DataFrame(nuevas_reps)
        st.write("### 📋 Vista previa de los nuevos datos:")
        st.dataframe(df_nuevas.head(5))

        if es_id_valido and st.button("🚀 ENVIAR A DROPBOX CENTRAL"):
            with st.spinner("Conectando con la base de datos del grupo en Dropbox..."):
                try:
                    # 1. Descargar el Excel actual de Dropbox
                    _, res = dbx.files_download(DBX_PATH)
                    df_historial = pd.read_excel(io.BytesIO(res.content))
                    
                    # 2. Unir datos viejos con nuevos
                    df_final = pd.concat([df_historial, df_nuevas], ignore_index=True)
                    
                    # 3. Darle estilo y prepararlo en memoria
                    buffer = io.BytesIO()
                    estilizar_excel(df_final, buffer)
                    
                    # 4. Volver a subirlo, sobreescribiendo el viejo
                    dbx.files_upload(buffer.getvalue(), DBX_PATH, mode=dropbox.files.WriteMode.overwrite)
                    
                    st.success("✅ ¡Datos guardados exitosamente en Dropbox!")
                    st.balloons()
                    
                except Exception as e:
                    st.error(f"❌ Error al conectar con Dropbox. Detalle: {e}")
